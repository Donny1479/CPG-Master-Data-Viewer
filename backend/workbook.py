from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from .schema import DIMENSION_KEYS, EXCEL_TO_APPWRITE_COLUMN, METRIC_KEYS, REQUIRED_MASTER_HEADERS

MASTER_SHEET = "Master Table"


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def import_run_id_for_file(path: str | Path) -> str:
    return f"imp_{file_sha256(path)[:28]}"


def appwrite_row_id(row: dict[str, Any]) -> str:
    stable = {
        "import_run_id": row.get("import_run_id"),
        "source_sheet": row.get("source_sheet"),
        "source_pull_type": row.get("source_pull_type"),
        "source_row_number": row.get("source_row_number"),
    }
    digest = hashlib.sha1(json.dumps(stable, sort_keys=True).encode("utf-8")).hexdigest()
    return f"r{digest[:31]}"


def row_hash(row: dict[str, Any]) -> str:
    clean = {key: value for key, value in row.items() if key != "$id"}
    return hashlib.sha256(json.dumps(clean, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)

    text = str(value).strip()
    if not text or text in {"-", "--", "NA", "N/A"}:
        return None

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace("$", "").replace(",", "").strip()
    if text.endswith("%"):
        text = text[:-1].strip()

    if not re.fullmatch(r"[-+]?\d*\.?\d+", text):
        return None

    number = float(text)
    return -number if negative else number


def normalize_cell(appwrite_key: str, value: Any) -> Any:
    if appwrite_key in METRIC_KEYS:
        return parse_number(value)

    if value is None:
        return None
    text = str(value).strip()
    return text or None


def master_headers(path: str | Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if MASTER_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain required sheet: {MASTER_SHEET}")
    worksheet = workbook[MASTER_SHEET]
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(value).strip() if value is not None else "" for value in first_row]


def validate_master_headers(headers: list[str]) -> None:
    missing = [header for header in REQUIRED_MASTER_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Master Table is missing required headers: {missing}")


def iter_master_rows(
    path: str | Path,
    import_run_id: str | None = None,
    limit: int | None = None,
) -> Iterator[dict[str, Any]]:
    source_path = Path(path)
    source_sha = file_sha256(source_path)
    import_run_id = import_run_id or f"imp_{source_sha[:28]}"

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if MASTER_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain required sheet: {MASTER_SHEET}")

    worksheet = workbook[MASTER_SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    validate_master_headers(headers)

    emitted = 0
    for source_row_number, values in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue

        output: dict[str, Any] = {
            "source_sheet": MASTER_SHEET,
            "source_pull_type": "master_table",
            "source_file": source_path.name,
            "source_file_sha256": source_sha,
            "import_run_id": import_run_id,
            "source_row_number": source_row_number,
        }

        for header, value in zip(headers, values):
            key = EXCEL_TO_APPWRITE_COLUMN.get(header)
            if key is None:
                continue
            normalized = normalize_cell(key, value)
            if normalized is not None:
                output[key] = normalized

        output["row_hash"] = row_hash(output)
        output["$id"] = appwrite_row_id(output)
        yield output

        emitted += 1
        if limit is not None and emitted >= limit:
            return


def summarize_workbook(path: str | Path) -> dict[str, Any]:
    source_path = Path(path)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    summary: dict[str, Any] = {
        "file": str(source_path),
        "size_mb": round(source_path.stat().st_size / 1024 / 1024, 2),
        "sheets": [],
    }

    for worksheet in workbook.worksheets:
        summary["sheets"].append(
            {
                "name": worksheet.title,
                "rows": worksheet.max_row,
                "columns": worksheet.max_column,
            }
        )

    rows = list(iter_master_rows(source_path))
    counters = {key: Counter() for key in DIMENSION_KEYS if key not in {"source_file_sha256", "row_hash"}}
    for row in rows:
        for key in counters:
            value = row.get(key)
            if value is not None:
                counters[key][str(value)] += 1

    summary["master_table"] = {
        "rows": len(rows),
        "columns": len(master_headers(source_path)),
        "distincts": {
            key: {"count": len(counter), "top": counter.most_common(8)}
            for key, counter in counters.items()
            if counter
        },
    }
    return summary
