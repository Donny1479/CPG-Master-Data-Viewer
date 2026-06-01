from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from .schema import EXCEL_TO_APPWRITE_COLUMN
from .workbook import appwrite_row_id, file_sha256, normalize_cell, row_hash

SOURCE_PULL_TYPES = {
    "topline_brands",
    "single_serve",
    "instant",
    "rg",
}

SKIP_SHEET_NAMES = {
    "instructions_disclaimers",
    "detailed instructions",
    "glossary",
    "dashboard",
    "summary_coffee category",
    "coffee category",
    "summary_coffee customer",
    "summary_coffee customer ",
    "coffee customer",
    "master table",
    "index",
}


@dataclass(frozen=True)
class SourceSheet:
    pull_type: str
    file_path: Path
    sheet_name: str
    header_row: int
    headers: list[str]


def combined_import_run_id(paths: list[str | Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(Path(item).resolve() for item in paths):
        digest.update(path.name.encode("utf-8"))
        digest.update(file_sha256(path).encode("utf-8"))
    return f"imp_{digest.hexdigest()[:28]}"


def classify_headers(headers: list[str]) -> str | None:
    header_set = {header for header in headers if header}
    if (
        {"Markets", "Periods", "Products"}.issubset(header_set)
        and "TH CATEGORY" not in header_set
        and "BRAND" not in header_set
    ):
        return "topline_brands"
    if {"Markets", "Periods", "NUMBER OF PACKETS", "TH CONTAINER TYPE"}.issubset(header_set):
        return "single_serve"
    if {"Markets", "Periods", "SIZE", "TH CONTAINER TYPE"}.issubset(header_set):
        return "rg"
    if {"Markets", "Periods", "SIZE", "TH CATEGORY", "BRAND"}.issubset(header_set):
        return "instant"
    return None


def discover_source_sheets(paths: list[str | Path]) -> dict[str, SourceSheet]:
    discovered: dict[str, SourceSheet] = {}

    for raw_path in paths:
        path = Path(raw_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            if worksheet.title.strip().lower() in SKIP_SHEET_NAMES:
                continue
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=min(25, worksheet.max_row), values_only=True),
                start=1,
            ):
                headers = [str(value).strip() if value is not None else "" for value in row]
                pull_type = classify_headers(headers)
                if not pull_type:
                    continue
                if pull_type in discovered:
                    other = discovered[pull_type]
                    raise ValueError(
                        f"Duplicate Nielsen pull type '{pull_type}' found in "
                        f"{other.file_path.name}/{other.sheet_name} and {path.name}/{worksheet.title}"
                    )
                discovered[pull_type] = SourceSheet(pull_type, path, worksheet.title, row_number, headers)
                break

    missing = sorted(SOURCE_PULL_TYPES - set(discovered))
    if missing:
        found = ", ".join(sorted(discovered)) or "none"
        raise ValueError(f"Missing Nielsen pull(s): {missing}. Found: {found}")

    return discovered


def product_case(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    fixed = text.title()
    fixed = fixed.replace("Ao", "AO")
    fixed = fixed.replace("Mccafe", "McCafe")
    return fixed


def build_product(brand: Any, th_brand: Any, size: Any) -> str | None:
    brand_raw = "" if brand is None else str(brand).strip()
    th_brand_raw = "" if th_brand is None else str(th_brand).strip()
    size_raw = "" if size is None else str(size).strip()

    brand_text = (brand_raw or th_brand_raw).upper().strip()
    size_text = size_raw.upper().strip()

    adjusted_brand = brand_text
    if "MAXWELL HOUSETASSIMO" in brand_text and "TASSIMO" in size_text:
        adjusted_brand = brand_text.replace("TASSIMO", "").strip()

    brand_no_space = adjusted_brand.replace(" ", "")
    size_words = [word.strip() for word in size_text.split(" ") if word.strip()]
    filtered_size_words = [
        word for word in size_words if word.replace(" ", "") not in brand_no_space
    ]
    cleaned_size = " ".join(filtered_size_words).strip()

    if adjusted_brand:
        product = f"{adjusted_brand} {cleaned_size}".strip()
    else:
        product = cleaned_size
    return product_case(product)


def _source_file_shas(source_sheets: dict[str, SourceSheet]) -> dict[Path, str]:
    return {sheet.file_path: file_sha256(sheet.file_path) for sheet in source_sheets.values()}


def iter_compiled_pull_rows(
    paths: list[str | Path],
    import_run_id: str | None = None,
    limit: int | None = None,
) -> Iterator[dict[str, Any]]:
    source_sheets = discover_source_sheets(paths)
    source_shas = _source_file_shas(source_sheets)
    import_run_id = import_run_id or combined_import_run_id(paths)

    emitted = 0
    for pull_type in ["topline_brands", "single_serve", "instant", "rg"]:
        source = source_sheets[pull_type]
        workbook = load_workbook(source.file_path, read_only=True, data_only=True)
        worksheet = workbook[source.sheet_name]
        rows = worksheet.iter_rows(min_row=source.header_row, values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]

        for source_row_number, values in enumerate(rows, start=source.header_row + 1):
            raw = dict(zip(headers, values))
            if pull_type == "topline_brands":
                product = product_case(raw.get("Products"))
                size = raw.get("SIZE")
            else:
                size = raw.get("SIZE", raw.get("NUMBER OF PACKETS"))
                product = build_product(raw.get("BRAND"), raw.get("TH BRAND"), size)

            market = normalize_cell("market", raw.get("Markets"))
            period = normalize_cell("period", raw.get("Periods"))
            if not market or not period or not product:
                continue

            master_like: dict[str, Any] = dict(raw)
            master_like["Products"] = product
            master_like["SIZE"] = size

            output: dict[str, Any] = {
                "source_sheet": source.sheet_name,
                "source_pull_type": pull_type,
                "source_file": source.file_path.name,
                "source_file_sha256": source_shas[source.file_path],
                "import_run_id": import_run_id,
                "source_row_number": source_row_number,
            }

            for excel_header, appwrite_key in EXCEL_TO_APPWRITE_COLUMN.items():
                normalized = normalize_cell(appwrite_key, master_like.get(excel_header))
                if normalized is not None:
                    output[appwrite_key] = normalized

            output["row_hash"] = row_hash(output)
            output["$id"] = appwrite_row_id(output)
            yield output

            emitted += 1
            if limit is not None and emitted >= limit:
                return


def summarize_source_pulls(paths: list[str | Path]) -> dict[str, Any]:
    source_sheets = discover_source_sheets(paths)
    rows = list(iter_compiled_pull_rows(paths))
    counts: dict[str, int] = {}
    for row in rows:
        pull_type = row["source_pull_type"]
        counts[pull_type] = counts.get(pull_type, 0) + 1

    return {
        "import_run_id": combined_import_run_id(paths),
        "sources": {
            pull_type: {
                "file": str(sheet.file_path),
                "sheet": sheet.sheet_name,
                "header_row": sheet.header_row,
                "rows": counts.get(pull_type, 0),
            }
            for pull_type, sheet in sorted(source_sheets.items())
        },
        "total_rows": len(rows),
        "sample_row": rows[0] if rows else None,
    }


def row_fingerprint(row: dict[str, Any]) -> str:
    return hashlib.sha256(json.dumps(row, sort_keys=True, default=str).encode("utf-8")).hexdigest()
